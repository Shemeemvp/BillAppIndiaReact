from django.shortcuts import render, redirect
from django.contrib.auth.models import User, auth
from django.contrib import messages
from random import randint
from django.core.mail import send_mail, EmailMessage
from io import BytesIO
from django.db import transaction
from django.conf import settings
from django.db import connection
from datetime import datetime, date, timedelta

# import requests
from decimal import Decimal
from num2words import num2words
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from .models import *
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Protection, Alignment
from django.db.models import Q
from rest_framework_simplejwt.tokens import RefreshToken
from rest_framework_simplejwt.views import TokenObtainPairView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser
import json
from rest_framework.decorators import api_view, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework import generics, status
from .serializers import *
from copy import deepcopy
from collections import defaultdict
from django.db.models import Sum

# Create your views here.


def home(request):
    return HttpResponse("Okay")


@api_view(("POST",))
def registerTrialUser(request):
    try:
        usrnm = request.data["username"]
        eml = request.data["email"]
        phn = request.data["phone_number"]
        cmpny = request.data["company_name"]
        pswrd = request.data["password"]
        cpswrd = request.data["confirm_password"]

        if User.objects.filter(username=usrnm).exists():
            return Response(
                {
                    "status": False,
                    "message": "This username already exists. Try another..",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        elif User.objects.filter(email=eml).exists():
            return Response(
                {
                    "status": False,
                    "message": "Email already exists. Try another..",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        elif Company.objects.filter(phone_number=phn).exists():
            return Response(
                {
                    "status": False,
                    "message": "Phone number already exists. Try another..",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        elif Company.objects.filter(company_name__iexact=cmpny.lower()).exists():
            return Response(
                {
                    "status": False,
                    "message": "Company Name already exists. Try another..",
                },
                status=status.HTTP_400_BAD_REQUEST,
            )
        else:
            if pswrd == cpswrd:
                userInfo = User.objects.create_user(
                    username=usrnm,
                    email=eml,
                    password=pswrd,
                )
                # cData = User.objects.get(id=userInfo.id)
                request.data["user"] =  userInfo.id
                serializer = CompanySerializer(data=request.data)
                if serializer.is_valid():
                    serializer.save()
                    cmp = Company.objects.get(cmp_id=serializer.data["cmp_id"])

                    # storing trial data
                    start = date.today()
                    end = start + timedelta(days=30)
                    trial = ClientTrials(
                        user=userInfo,
                        company=cmp,
                        start_date=start,
                        end_date=end,
                        trial_status=True,
                        purchase_start_date=None,
                        purchase_end_date=None,
                        purchase_status="null",
                        payment_term=None,
                        subscribe_status="null",
                    )
                    trial.save()

                    return Response(
                        {"status": True, "data": serializer.data},
                        status=status.HTTP_201_CREATED,
                    )
                else:
                    return Response(
                        {"status": False, "data": serializer.errors},
                        status=status.HTTP_400_BAD_REQUEST,
                    )
            else:
                return Response(
                    {
                        "status": False,
                        "message": "Password and confirm password does not match..",
                    },
                    status=status.HTTP_400_BAD_REQUEST,
                )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def userLogin(request):
    try:
        uName = request.data["username"]
        password = request.data["password"]

        log_user = auth.authenticate(username=uName, password=password)
        if log_user is not None:
            user = User.objects.get(username=uName)
            if user.is_staff:
                auth.login(request, log_user)

                refresh = RefreshToken.for_user(user)

                return Response(
                    {
                        "status": True,
                        "is_staff": True,
                        "user": str(user.id),
                        "refresh": str(refresh),
                        "access": str(refresh.access_token),
                        "role": "Admin",
                    },
                    status=status.HTTP_200_OK,
                )
            else:
                refresh = RefreshToken.for_user(user)
                log_status = ClientTrials.objects.get(user=user.id)

                if (
                    log_status.purchase_status == "null"
                    and log_status.trial_status == True
                ):
                    exp_days = (log_status.end_date - date.today()).days
                    if exp_days < 0:
                        log_status.trial_status = False
                        log_status.save()
                        pass
                elif log_status.purchase_status == "valid":
                    sub_exp_days = (log_status.purchase_end_date - date.today()).days
                    if sub_exp_days < 0:
                        log_status.purchase_status = "expired"
                        log_status.save()
                        pass
                else:
                    pass

                if log_status.purchase_status == "valid":
                    auth.login(request, user)
                    return Response(
                        {
                            "status": True,
                            "is_staff": False,
                            "user": str(user.id),
                            "refresh": str(refresh),
                            "access": str(refresh.access_token),
                            "role": "User",
                        },
                        status=status.HTTP_200_OK,
                    )
                elif log_status.purchase_status == "expired":
                    return Response(
                        {
                            "status": False,
                            "message": "Your Subscription has been expired.! Contact Admin.",
                            "is_staff": False,
                            "user": str(user.id),
                            "refresh": str(refresh),
                            "access": str(refresh.access_token),
                        },
                        status=status.HTTP_200_OK,
                    )
                elif log_status.purchase_status == "cancelled":
                    return Response(
                        {
                            "status": False,
                            "message": "Your Subscription has been Cancelled.! Contact Admin.",
                            "is_staff": False,
                            "user": str(user.id),
                            "refresh": str(refresh),
                            "access": str(refresh.access_token),
                        },
                        status=status.HTTP_200_OK,
                    )
                else:
                    if log_status.trial_status:
                        auth.login(request, user)
                        return Response(
                            {
                                "status": True,
                                "is_staff": False,
                                "user": str(user.id),
                                "refresh": str(refresh),
                                "access": str(refresh.access_token),
                                "role": "User",
                            },
                            status=status.HTTP_200_OK,
                        )
                    else:
                        return Response(
                            {
                                "status": False,
                                "message": "Your Trial period has been expired.! Contact Admin.",
                                "is_staff": False,
                                "user": str(user.id),
                                "refresh": str(refresh),
                                "access": str(refresh.access_token),
                            },
                            status=status.HTTP_200_OK,
                        )
        else:
            return Response(
                {"status": False, "message": "Invalid username or password, try again"},
                status=status.HTTP_404_NOT_FOUND,
            )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def validateEmail(request):
    email = request.GET["email"]

    if User.objects.filter(email=email).exists():
        return JsonResponse({"is_taken": True})
    else:
        return JsonResponse({"is_taken": False})


@api_view(("GET",))
def validateUsername(request):
    uName = request.GET["user"]

    if User.objects.filter(username=uName).exists():
        return JsonResponse({"is_taken": True})
    else:
        return JsonResponse({"is_taken": False})


@api_view(("GET",))
def validatePhone(request):
    number = request.GET["phone"]

    if Company.objects.filter(phone_number=number).exists():
        return JsonResponse({"is_taken": True})
    else:
        return JsonResponse({"is_taken": False})


@api_view(("GET",))
def validateCompany(request):
    cmp = request.GET["company"]

    if Company.objects.filter(company_name__iexact=cmp.lower()).exists():
        return JsonResponse({"is_taken": True})
    else:
        return JsonResponse({"is_taken": False})


# Admin


@api_view(("GET",))
def fetchRegisteredClients(request):
    try:
        # Checking clients Subscription and trial status, updates if expired.
        trials = ClientTrials.objects.all()
        for cStatus in trials:
            if cStatus.purchase_status == "null" and cStatus.trial_status == True:
                exp_days = (cStatus.end_date - date.today()).days
                if exp_days < 0:
                    cStatus.trial_status = False
                    cStatus.save()
                    pass
            elif cStatus.purchase_status == "valid":
                sub_exp_days = (cStatus.purchase_end_date - date.today()).days
                if sub_exp_days < 0:
                    cStatus.purchase_status = "expired"
                    cStatus.save()
                    pass
            else:
                pass

        all_companies = Company.objects.all()
        clients = []
        for i in all_companies:
            if ClientTrials.objects.filter(company=i).exists():
                trial = ClientTrials.objects.filter(company=i).first()
            else:
                trial = None

            dict = {
                "user_id": i.user.id,
                "company_name": i.company_name,
                "email": i.user.email,
                "contact": i.phone_number,
                "gstin": i.gst_number,
                "start_date": trial.start_date if trial else "",
            }
            clients.append(dict)

        return Response(
            {"status": True, "clients": clients},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("DELETE",))
def deleteClient(request, id):
    try:
        user = User.objects.get(id=id)
        user.delete()
        return Response(
            {"status": True},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def fetchPaymentTerms(request):
    try:
        terms = PaymentTerms.objects.all()
        termsSerializer = PaymentTermsSerializer(terms, many=True)
        return Response(
            {"status": True, "terms": termsSerializer.data},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("DELETE",))
def deletePaymentTerm(request, id):
    try:
        term = PaymentTerms.objects.get(id=id)
        term.delete()
        return Response(
            {"status": True},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def createNewPaymentTerm(request):
    try:
        dur = request.data["duration"]
        term = request.data["term"]
        dys = int(dur) if term == "Days" else int(dur) * 30
        request.data["days"] = dys

        # PaymentTerms.objects.create(duration=dur, term=term, days=dys)
        serializer = PaymentTermsSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            return Response(
                {"status": True, "data": serializer.data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"status": False, "data": serializer.errors},
                status=status.HTTP_200_OK,
            )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def fetchDemoClients(request):
    try:
        d_clients = ClientTrials.objects.filter(
            Q(trial_status=True) | (Q(trial_status=False) & Q(subscribe_status="yes"))
        )
        clients = []
        for i in d_clients:

            dict = {
                "trial_id": i.id,
                "company_name": i.company.company_name,
                "email": i.user.email,
                "contact": i.company.phone_number,
                "gstin": i.company.gst_number,
                "end_date": i.end_date,
                "purchase_status": (
                    "Interested"
                    if i.subscribe_status == "yes"
                    else "Not Interested" if i.subscribe_status == "no" else "N/A"
                ),
            }
            clients.append(dict)

        return Response(
            {"status": True, "clients": clients},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def clientPurchase(request):
    try:
        client = ClientTrials.objects.get(id=request.data["id"])
        start = request.data["purchaseDate"]
        end = request.data["endDate"]
        term = PaymentTerms.objects.get(id=request.data["term"])

        client.purchase_start_date = start
        client.purchase_end_date = end
        client.payment_term = str(term.duration) + " " + term.term
        client.purchase_status = "valid"
        client.trial_status = False
        client.subscribe_status = "purchased"
        client.save()

        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def fetchPurchasedClients(request):
    try:
        d_clients = ClientTrials.objects.exclude(purchase_status="null")
        clients = []
        for i in d_clients:

            dict = {
                "trial_id": i.id,
                "company_name": i.company.company_name,
                "email": i.user.email,
                "contact": i.company.phone_number,
                "gstin": i.company.gst_number,
                "start_date": i.purchase_start_date,
                "end_date": i.purchase_end_date,
                "payment_term": i.payment_term,
                "purchase_status": i.purchase_status,
            }
            clients.append(dict)

        return Response(
            {"status": True, "clients": clients},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def cancelSubscription(request):
    try:
        client = ClientTrials.objects.get(id=request.data["id"])
        client.purchase_status = "cancelled"
        client.save()

        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getAdminNotifications(request):
    try:
        renewals = ClientTrials.objects.filter(subscribe_status="yes")
        serializer = TrialSerializer(renewals, many=True)
        if renewals:
            return Response(
                {"status": True, "renewals": serializer.data}, status=status.HTTP_200_OK
            )
        else:
            return Response({"status": False}, status=status.HTTP_404_NOT_FOUND)
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# User


@api_view(("GET",))
def getSelfData(request, id):
    try:
        user = User.objects.get(id=id)
        img = None
        name = None
        date = None

        status = ClientTrials.objects.get(user=user)
        if status.purchase_status == "valid":
            date = status.purchase_end_date

        if user:
            usrData = Company.objects.get(user=user)
            img = usrData.logo.url if usrData.logo else None
            name = usrData.company_name
        else:
            usrData = None

        cmpSerializer = CompanySerializer(usrData)
        details = {
            "name": name,
            "image": img,
            "user": UserSerializer(user).data,
            "endDate": date,
            "company": cmpSerializer.data,
        }

        return Response({"status": True, "data": details})
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getDashboardDetails(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        sales = Sales.objects.filter(cid=cmp)
        purchases = Purchases.objects.filter(cid=cmp)
        items = Items.objects.filter(cid=cmp)
        todSale = 0
        totSale = 0
        todPurchase = 0
        totPurchase = 0
        for i in sales:
            totSale += i.total_amount
            if i.date == date.today():
                todSale += i.total_amount

        for i in purchases:
            totPurchase += i.total_amount
            if i.date == date.today():
                todPurchase += i.total_amount

        # Chart data

        data1 = []
        data2 = []
        data3 = []
        data4 = []
        data5 = []
        label = []

        for yr in range((date.today().year) - 4, (date.today().year) + 1):
            label.append(yr)
            salesAmount = 0
            purchaseAmount = 0
            for i in sales:
                if i.date.year == yr:
                    salesAmount += i.total_amount

            for i in purchases:
                if i.date.year == yr:
                    purchaseAmount += i.total_amount

            data1.append(float(salesAmount))
            data2.append(float(purchaseAmount))

            stockIn = 0
            stockOut = 0
            stockBalance = 0
            for i in Item_transactions.objects.filter(cid=cmp).filter(type="Purchase"):
                if i.date.year == yr:
                    stockIn += i.quantity

            for i in Item_transactions.objects.filter(cid=cmp).filter(type="Sale"):
                if i.date.year == yr:
                    stockOut += i.quantity

            for i in Item_transactions.objects.filter(cid=cmp):
                if i.date.year == yr and (
                    i.type == "Opening Stock"
                    or i.type == "Add Stock"
                    or i.type == "Purchase"
                ):
                    stockBalance += i.quantity

                if i.date.year == yr and (i.type == "Reduce Stock" or i.type == "Sale"):
                    stockBalance -= i.quantity

            data3.append(stockIn)
            data4.append(stockOut)
            data5.append(stockBalance)

        context = {
            "todSale": f"{todSale:.2f}",
            "totSale": f"{totSale:.2f}",
            "todPurchase": f"{todPurchase:.2f}",
            "totPurchase": f"{totPurchase:.2f}",
            "salesData": data1,
            "purchaseData": data2,
            "stockIn": data3,
            "stockOut": data4,
            "stockBalance": data5,
            "label": label,
        }
        return Response({"status": True, "data": context}, status=status.HTTP_200_OK)
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def endDate(request, id):
    try:
        usr = User.objects.get(id=id)

        date = ""
        cStatus = ClientTrials.objects.get(user=usr)
        if cStatus.purchase_status == "valid":
            date = cStatus.purchase_end_date

        return Response({"status": True, "endDate": date}, status=status.HTTP_200_OK)
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def fetchNotifications(request, id):
    try:
        usr = User.objects.get(id=id)
        cStatus = ClientTrials.objects.get(user=usr)

        if (
            cStatus.purchase_status == "null"
            and cStatus.trial_status == True
            and cStatus.subscribe_status != "yes"
        ):
            exp_days = (cStatus.end_date - date.today()).days
            if exp_days <= 10:
                return Response(
                    {"status": True, "days": exp_days, "subscribe": False},
                    status=status.HTTP_200_OK,
                )
            else:
                return Response({"status": False}, status=status.HTTP_200_OK)
        elif (
            cStatus.purchase_status == "null"
            and cStatus.trial_status == True
            and cStatus.subscribe_status == "yes"
        ):
            exp_days = (cStatus.end_date - date.today()).days
            if exp_days <= 10:
                return Response(
                    {"status": True, "days": exp_days, "subscribe": True},
                    status=status.HTTP_200_OK,
                )
            else:
                return Response({"status": False}, status=status.HTTP_200_OK)
        else:
            return Response({"status": False}, status=status.HTTP_200_OK)
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def changeSubscribeStatus(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        trial = ClientTrials.objects.get(user=usr)

        trial.subscribe_status = request.data["status"]
        trial.save()
        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("PUT",))
@parser_classes((MultiPartParser, FormParser))
def updateCompanyLogo(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)

        serializer = CompanySerializer(cmp, data=request.data, partial=True)
        if serializer.is_valid():
            serializer.save()

            return Response(
                {"status": True, "data": serializer.data}, status=status.HTTP_200_OK
            )
        else:
            return Response(
                {"status": False, "data": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def removeCompanyLogo(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)

        cmp.logo = None
        cmp.save()

        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def updateProfileData(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)

        uName = request.data["username"]
        email = request.data["email"]
        if uName != usr.username and User.objects.filter(username=uName).exists():
            return Response(
                {"status": False, "message": "Username exists.!"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        else:
            usr.username = uName
            usr.save()

        if email != usr.email and User.objects.filter(email=email).exists():
            return Response(
                {"status": False, "message": "Email exists.!"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        else:
            usr.email = email
            usr.save()

        cmpName = request.data["company_name"]
        cmpPhone = request.data["phone_number"]
        cmpGst = request.data["gst_number"]
        cmpAddress = request.data["address"]
        cmpState = request.data["state"]
        cmpCountry = request.data["country"]

        cmp.company_name = cmpName
        cmp.phone_number = cmpPhone
        cmp.gst_number = cmpGst
        cmp.address = cmpAddress
        cmp.state = cmpState
        cmp.country = cmpCountry

        cmp.save()

        return Response(
            {"status": True},
            status=status.HTTP_200_OK,
        )

    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getItemUnits(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        units = Item_units.objects.filter(cid=cmp)
        serializer = UnitsSerializer(units, many=True)
        return Response(
            {"status": True, "units": serializer.data}, status=status.HTTP_200_OK
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def createNewUnit(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)
        sym = request.data["symbol"]
        name = request.data["name"]
        unit = Item_units(cid=cmp, symbol=sym, name=name)
        unit.save()
        return Response(
            {"status": True},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def checkItemBarcode(request):
    try:
        usr = User.objects.get(id=request.GET["Id"])
        cmp = Company.objects.get(user=usr)

        bc = request.GET["barcode"]
        if Items.objects.filter(cid=cmp, barcode=bc).exists():
            item = Items.objects.get(cid=cmp, barcode=bc)
            return Response(
                {
                    "status": False,
                    "message": f"Barcode is already associated with item - '{item.name}',\nTry again..",
                },
                status=status.HTTP_226_IM_USED,
            )
        else:
            return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def createNewItem(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)
        request.data["cid"] = cmp.cmp_id

        bc = request.data["barcode"]
        if bc != "":
            bc = bc.upper()
        if bc != "" and Items.objects.filter(cid=cmp, barcode=bc).exists():
            return Response(
                {"status": False, "message": "Barcode already exists, try another.!"},
                status=status.HTTP_226_IM_USED,
            )
        tax = request.data["tax_reference"]
        request.data["tax"] = "Taxable" if tax else "Non-taxable"
        request.data["gst"] = "GST0[0%]" if not tax else request.data["gst"]
        request.data["igst"] = "IGST0[0%]" if not tax else request.data["igst"]

        serializer = ItemSerializer(data=request.data)
        if serializer.is_valid():
            serializer.save()
            item = Items.objects.get(id=serializer.data["id"])

            transaction = Item_transactions(
                cid=cmp,
                item=item,
                type="Opening Stock",
                date=item.date,
                quantity=item.stock,
            )
            transaction.save()

            return Response(
                {"status": True, "data": serializer.data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"status": False, "data": serializer.errors},
                status=status.HTTP_200_OK,
            )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getItems(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        itms = Items.objects.filter(cid=cmp)
        iData = Items.objects.filter(cid=cmp).first()
        trans = Item_transactions.objects.filter(cid=cmp, item=iData)

        serializer = ItemSerializer(itms, many=True)
        itemSerializer = ItemSerializer(iData)
        trnsSerializer = ItemTransSerializer(trans, many=True)
        return Response(
            {
                "status": True,
                "items": serializer.data,
                "firstItem": itemSerializer.data,
                "transactions": trnsSerializer.data,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getItemDetails(request):
    try:
        iData = Items.objects.get(id=request.GET["itemId"])
        trans = Item_transactions.objects.filter(item=iData)

        trns = Item_transactions.objects.get(item=iData, type="Opening Stock")
        op_stock = trns.quantity

        itemSerializer = ItemSerializer(iData)
        trnsSerializer = ItemTransSerializer(trans, many=True)
        return Response(
            {
                "status": True,
                "firstItem": itemSerializer.data,
                "transactions": trnsSerializer.data,
                "op_stock": op_stock,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def updateStock(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)

        item = Items.objects.get(cid=cmp, id=request.data["item_id"])
        adj = request.data["adjust"]
        if adj:
            item.stock += int(request.data["stock"])
            item.save()

            trns = Item_transactions(
                cid=cmp,
                item=item,
                type="Add Stock",
                date=request.data["date"],
                quantity=request.data["stock"],
            )
            trns.save()
        else:
            item.stock -= int(request.data["stock"])
            item.save()

            trns = Item_transactions(
                cid=cmp,
                item=item,
                type="Reduce Stock",
                date=request.data["date"],
                quantity=request.data["stock"],
            )
            trns.save()
        return Response(
            {"status": True},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getTransactionDetails(request, id):
    try:
        trans = Item_transactions.objects.get(id=id)
        trnsSerializer = ItemTransSerializer(trans)
        return Response(
            {
                "status": True,
                "transaction": trnsSerializer.data,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("DELETE",))
def deleteTransaction(request, id):
    try:
        trns = Item_transactions.objects.get(id=id)
        item = Items.objects.get(id=trns.item.id)
        if trns.type == "Add Stock":
            item.stock -= trns.quantity
        elif trns.type == "Reduce Stock":
            item.stock += trns.quantity

        item.save()
        trns.delete()
        return Response(
            {
                "status": True,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def updateTransaction(request):
    try:
        trns = Item_transactions.objects.get(id=request.data["trans_id"])
        item = Items.objects.get(id=trns.item.id)
        crQty = trns.quantity
        chQty = int(request.data["quantity"])
        diff = abs(crQty - chQty)

        type = request.data["type"]
        if str(type).lower() == "reduce stock" and chQty > crQty:
            item.stock -= diff
        elif str(type).lower() == "reduce stock" and chQty < crQty:
            item.stock += diff
        elif str(type).lower() == "add stock" and chQty > crQty:
            item.stock += diff
        elif str(type).lower() == "add stock" and chQty < crQty:
            item.stock -= diff

        if str(type).lower() == "opening stock" and chQty > crQty:
            item.stock += diff
        elif str(type).lower() == "opening stock" and chQty < crQty:
            item.stock -= diff

        trns.quantity = request.data["quantity"]
        trns.date = request.data["date"]
        trns.save()
        item.save()

        return Response(
            {"status": True},
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("DELETE",))
def deleteItem(request, id):
    try:
        item = Items.objects.get(id=id)

        if (
            Sales_items.objects.filter(item=item).exists()
            or Purchase_items.objects.filter(item=item).exists()
        ):
            return Response(
                {
                    "status": False,
                    "message": f"Item cannot be deleted because of Sales or Purchase transactions exists for `{item.name}`.",
                },
                status=status.HTTP_200_OK,
            )
        else:
            item.delete()
            return Response(
                {
                    "status": True,
                },
                status=status.HTTP_200_OK,
            )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("PUT",))
def updateItem(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)
        request.data["cid"] = cmp.cmp_id

        item = Items.objects.get(id=request.data["item_id"])
        trns = (
            Item_transactions.objects.filter(item=item.id)
            .filter(type="Opening Stock")
            .first()
        )
        crQty = trns.quantity
        chQty = int(request.data["stock"])
        diff = abs(crQty - chQty)

        if diff != 0 and chQty > crQty:
            request.data["stock"] = item.stock + diff
        elif diff != 0 and chQty < crQty:
            request.data["stock"] = item.stock - diff
        else:
            request.data["stock"] = item.stock

        bc = request.data["barcode"]
        if bc != "":
            bc = bc.upper()
        if (
            item.barcode != bc
            and bc != ""
            and Items.objects.filter(cid=cmp, barcode=bc).exists()
        ):
            return Response(
                {"status": False, "message": "Barcode already exists, try another.!"},
                status=status.HTTP_226_IM_USED,
            )
        tax = request.data["tax_reference"]
        request.data["tax"] = "Taxable" if tax else "Non-taxable"
        request.data["gst"] = "GST0[0%]" if not tax else request.data["gst"]
        request.data["igst"] = "IGST0[0%]" if not tax else request.data["igst"]

        serializer = ItemSerializer(item, data=request.data)
        if serializer.is_valid():
            serializer.save()

            trns.quantity = chQty
            trns.save()

            return Response(
                {"status": True, "data": serializer.data},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"status": False, "data": serializer.errors},
                status=status.HTTP_200_OK,
            )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# Sales
@api_view(("GET",))
def fetchSalesData(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        items = Items.objects.filter(cid=cmp)

        itemSerializer = ItemSerializer(items, many=True)

        # Fetching last bill and assigning upcoming bill no as current + 1
        # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
        latest_bill = Sales.objects.filter(cid=cmp).order_by("-bill_no").first()

        if latest_bill:
            last_number = int(latest_bill.bill_number)
            new_number = last_number + 1
        else:
            new_number = 1

        if DeletedSales.objects.filter(cid=cmp).exists():
            deleted = DeletedSales.objects.get(cid=cmp)

            if deleted:
                while int(deleted.bill_number) >= new_number:
                    new_number += 1

        return Response(
            {
                "status": True,
                "items": itemSerializer.data,
                "billNo": new_number,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getItemData(request):
    try:
        usr = User.objects.get(id=request.GET["Id"])
        cmp = Company.objects.get(user=usr)
        id = request.GET["item"]

        item = Items.objects.get(id=id)
        data = {
            "hsn": item.hsn,
            "pur_rate": item.purchase_price,
            "sale_rate": item.sale_price,
            "tax": True if item.tax == "Taxable" else False,
            "gst": item.gst,
            "igst": item.igst,
        }
        return Response({"status": True, "itemData": data}, status=status.HTTP_200_OK)
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getBarcodeDetails(request):
    try:
        usr = User.objects.get(id=request.GET["Id"])
        cmp = Company.objects.get(user=usr)
        bc = request.GET["barcode"]

        if Items.objects.filter(cid=cmp, barcode=bc).exists():
            item = Items.objects.get(cid=cmp, barcode=bc)
            hsn = item.hsn
            pur_rate = item.purchase_price
            sale_rate = item.sale_price
            tax = True if item.tax == "Taxable" else False
            gst = item.gst
            igst = item.igst
            data = {
                "id": item.id,
                "name": item.name,
                "hsn": hsn,
                "pur_rate": pur_rate,
                "sale_rate": sale_rate,
                "tax": tax,
                "gst": gst,
                "igst": igst,
            }
            return Response(
                {
                    "status": True,
                    "itemData": data,
                }
            )
        else:
            return Response({"status": False, "message": "No data found.!"})
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def createSales(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)

        mutable_data = deepcopy(request.data)
        mutable_data["cid"] = cmp.cmp_id

        serializer = SalesSerializer(data=mutable_data)
        if serializer.is_valid():
            serializer.save()

            salesItems = json.loads(request.data["salesItems"])
            sale = Sales.objects.get(bill_no=serializer.data["bill_no"])

            for ele in salesItems:
                itm = Items.objects.get(id=int(ele.get("item")))
                qty = int(ele.get("quantity"))
                hsn = ele.get("hsn")
                price = ele.get("price")
                tax = (
                    ele.get("taxGst")
                    if request.data["state_of_supply"] == "State"
                    else ele.get("taxIgst")
                )

                Sales_items.objects.create(
                    cid=cmp,
                    sid=sale,
                    item=itm,
                    name=itm.name,
                    hsn=hsn,
                    quantity=qty,
                    rate=float(price),
                    tax=tax,
                    total=float(ele.get("total")),
                )

            # Add sales details in items transactions
            for itm in salesItems:
                tItem = Items.objects.get(id=itm.get("item"), cid=cmp)
                Item_transactions.objects.create(
                    cid=cmp,
                    item=tItem,
                    type="Sale",
                    date=sale.date,
                    quantity=itm.get("quantity"),
                    bill_number=sale.bill_number,
                )
                tItem.stock -= int(itm.get("quantity"))
                tItem.save()

            return Response(
                {"status": True, "data": serializer.data}, status=status.HTTP_200_OK
            )
        else:
            return Response(
                {"status": False, "data": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getSalesBills(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        sales = Sales.objects.filter(cid=cmp)

        serializer = SalesSerializer(sales, many=True)
        return Response(
            {
                "status": True,
                "sales": serializer.data,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getSalesBillDetails(request):
    try:
        usr = User.objects.get(id=request.GET["Id"])
        cmp = Company.objects.get(user=usr)

        sale = Sales.objects.get(bill_no=request.GET["salesId"])
        items = Sales_items.objects.filter(sid=sale)

        saleSerializer = SalesSerializer(sale)
        saleItemsSerializer = SalesItemsSerializer(items, many=True)
        return Response(
            {
                "status": True,
                "bill": saleSerializer.data,
                "items": saleItemsSerializer.data,
                "cmp": {
                    "company_name": cmp.company_name,
                    "address": cmp.address,
                    "email": usr.email,
                    "state": cmp.state,
                    "country": cmp.country,
                    "phone_number": cmp.phone_number,
                    "gst_number": cmp.gst_number,
                },
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("DELETE",))
def deleteSalesBill(request, id):
    try:
        bill = Sales.objects.get(bill_no=id)
        cmp = bill.cid
        items = Sales_items.objects.filter(sid=bill)
        for i in items:
            itm = Items.objects.get(id=i.item.id)
            itm.stock += i.quantity
            itm.save()
            Item_transactions.objects.filter(
                bill_number=bill.bill_number, type="Sale", item=itm
            ).delete()
        Sales_items.objects.filter(sid=bill).delete()

        # Storing bill number to deleted table
        # if entry exists and lesser than the current, update and save => Only one entry per company

        if DeletedSales.objects.filter(cid=cmp).exists():
            deleted = DeletedSales.objects.get(cid=cmp)
            if deleted:
                if int(bill.bill_number) > int(deleted.bill_number):
                    deleted.bill_number = bill.bill_number
                    deleted.save()
        else:
            deleted = DeletedSales(cid=cmp, bill_number=bill.bill_number)
            deleted.save()

        bill.delete()
        return Response(
            {
                "status": True,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def salesBillPdf(request):
    try:
        id = request.GET["Id"]
        saleId = request.GET["saleId"]

        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        bill = Sales.objects.get(cid=cmp, bill_no=saleId)
        items = Sales_items.objects.filter(cid=cmp, sid=bill)

        total = bill.total_amount
        words_total = num2words(total)

        context = {"bill": bill, "cmp": cmp, "items": items, "total": words_total}

        template_path = "sales_bill_pdf.html"
        fname = "SalesBill_" + str(bill.bill_number)
        # Create a Django response object, and specify content_type as pdftemp_
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f"attachment; filename = {fname}.pdf"
        # find the template and render it.
        template = get_template(template_path)
        html = template.render(context)

        # create a pdf
        pisa_status = pisa.CreatePDF(html, dest=response)
        # if error then show some funny view
        if pisa_status.err:
            return HttpResponse("We had some errors <pre>" + html + "</pre>")
        return response
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def shareSalesBillToEmail(request):
    try:
        id = request.data["Id"]
        saleId = request.data["saleId"]

        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        bill = Sales.objects.get(cid=cmp, bill_no=saleId)
        items = Sales_items.objects.filter(cid=cmp, sid=bill)

        total = bill.total_amount
        words_total = num2words(total)

        emails_string = request.data["email_ids"]

        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(",")]
        email_message = request.data["email_message"]
        # print(emails_list)

        context = {"bill": bill, "cmp": cmp, "items": items, "total": words_total}

        template_path = "sales_bill_pdf.html"
        template = get_template(template_path)

        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        filename = f"SalesBill_{bill.bill_number}.pdf"
        subject = f"SalesBill_{bill.bill_number}"
        email = EmailMessage(
            subject,
            f"Hi,\nPlease find the attached details - SALES BILL-{bill.bill_number}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.phone_number}",
            from_email=settings.EMAIL_HOST_USER,
            to=emails_list,
        )
        email.attach(filename, pdf, "application/pdf")
        email.send(fail_silently=False)

        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("PUT",))
def updateSales(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)
        sale = Sales.objects.get(bill_no=request.data["saleId"])

        mutable_data = deepcopy(request.data)
        mutable_data["cid"] = cmp.cmp_id

        if request.data["party"] == "false":
            mutable_data["party_name"] = ""
            mutable_data["phone_number"] = ""
            mutable_data["gstin"] = ""

        serializer = SalesSerializer(sale, data=mutable_data)
        if serializer.is_valid():
            serializer.save()

            salesItems = json.loads(request.data["salesItems"])

            # delete old transactions and stock updates
            items = Sales_items.objects.filter(sid=sale)
            for i in items:
                itm = Items.objects.get(id=i.item.id)
                itm.stock += i.quantity
                itm.save()
                Item_transactions.objects.filter(
                    bill_number=sale.bill_number, type="Sale", item=itm
                ).delete()
            Sales_items.objects.filter(sid=sale).delete()

            # Add sales items and corresponding item transactions
            for ele in salesItems:
                itm = Items.objects.get(id=int(ele.get("item")))
                qty = int(ele.get("quantity"))
                hsn = ele.get("hsn")
                price = ele.get("price")
                tax = (
                    ele.get("taxGst")
                    if request.data["state_of_supply"] == "State"
                    else ele.get("taxIgst")
                )

                Sales_items.objects.create(
                    cid=cmp,
                    sid=sale,
                    item=itm,
                    name=itm.name,
                    hsn=hsn,
                    quantity=qty,
                    rate=float(price),
                    tax=tax,
                    total=float(ele.get("total")),
                )

            # Add sales details in items transactions
            for itm in salesItems:
                tItem = Items.objects.get(id=itm.get("item"), cid=cmp)
                Item_transactions.objects.create(
                    cid=cmp,
                    item=tItem,
                    type="Sale",
                    date=sale.date,
                    quantity=itm.get("quantity"),
                    bill_number=sale.bill_number,
                )
                tItem.stock -= int(itm.get("quantity"))
                tItem.save()

            return Response(
                {"status": True, "data": serializer.data}, status=status.HTTP_200_OK
            )
        else:
            return Response(
                {"status": False, "data": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# Purchase
@api_view(("GET",))
def fetchPurchaseData(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        items = Items.objects.filter(cid=cmp)

        itemSerializer = ItemSerializer(items, many=True)

        # Fetching last bill and assigning upcoming bill no as current + 1
        # Also check for if any bill is deleted and bill no is continuos w r t the deleted bill
        latest_bill = Purchases.objects.filter(cid=cmp).order_by("-bill_no").first()

        if latest_bill:
            last_number = int(latest_bill.bill_number)
            new_number = last_number + 1
        else:
            new_number = 1

        if DeletedPurchases.objects.filter(cid=cmp).exists():
            deleted = DeletedPurchases.objects.get(cid=cmp)

            if deleted:
                while int(deleted.bill_number) >= new_number:
                    new_number += 1

        return Response(
            {
                "status": True,
                "items": itemSerializer.data,
                "billNo": new_number,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def createPurchase(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)

        mutable_data = deepcopy(request.data)
        mutable_data["cid"] = cmp.cmp_id

        serializer = PurchaseSerializer(data=mutable_data)
        if serializer.is_valid():
            serializer.save()

            purchsItems = json.loads(request.data["salesItems"])
            purchase = Purchases.objects.get(bill_no=serializer.data["bill_no"])

            for ele in purchsItems:
                itm = Items.objects.get(id=int(ele.get("item")))
                qty = int(ele.get("quantity"))
                hsn = ele.get("hsn")
                price = ele.get("price")
                tax = (
                    ele.get("taxGst")
                    if request.data["state_of_supply"] == "State"
                    else ele.get("taxIgst")
                )

                Purchase_items.objects.create(
                    cid=cmp,
                    pid=purchase,
                    item=itm,
                    name=itm.name,
                    hsn=hsn,
                    quantity=qty,
                    rate=float(price),
                    tax=tax,
                    total=float(ele.get("total")),
                )

            # Add sales details in items transactions
            for itm in purchsItems:
                tItem = Items.objects.get(id=itm.get("item"), cid=cmp)
                Item_transactions.objects.create(
                    cid=cmp,
                    item=tItem,
                    type="Purchase",
                    date=purchase.date,
                    quantity=itm.get("quantity"),
                    bill_number=purchase.bill_number,
                )
                tItem.stock += int(itm.get("quantity"))
                tItem.save()

            return Response(
                {"status": True, "data": serializer.data}, status=status.HTTP_200_OK
            )
        else:
            return Response(
                {"status": False, "data": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getPurchaseBills(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        purchases = Purchases.objects.filter(cid=cmp)

        serializer = PurchaseSerializer(purchases, many=True)
        return Response(
            {
                "status": True,
                "sales": serializer.data,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getPurchaseBillDetails(request):
    try:
        usr = User.objects.get(id=request.GET["Id"])
        cmp = Company.objects.get(user=usr)

        purchase = Purchases.objects.get(bill_no=request.GET["purchaseId"])
        items = Purchase_items.objects.filter(pid=purchase)

        purchaseSerializer = PurchaseSerializer(purchase)
        purItemsSerializer = PurchaseItemsSerializer(items, many=True)
        return Response(
            {
                "status": True,
                "bill": purchaseSerializer.data,
                "items": purItemsSerializer.data,
                "cmp": {
                    "company_name": cmp.company_name,
                    "address": cmp.address,
                    "email": usr.email,
                    "state": cmp.state,
                    "country": cmp.country,
                    "phone_number": cmp.phone_number,
                    "gst_number": cmp.gst_number,
                },
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("DELETE",))
def deletePurchaseBill(request, id):
    try:
        bill = Purchases.objects.get(bill_no=id)
        cmp = bill.cid
        items = Purchase_items.objects.filter(pid=bill)
        for i in items:
            itm = Items.objects.get(id=i.item.id)
            itm.stock -= i.quantity
            itm.save()
            Item_transactions.objects.filter(
                bill_number=bill.bill_number, type="Purchase", item=itm
            ).delete()
        Purchase_items.objects.filter(pid=bill).delete()

        # Storing bill number to deleted table
        # if entry exists and lesser than the current, update and save => Only one entry per company

        if DeletedPurchases.objects.filter(cid=cmp).exists():
            deleted = DeletedPurchases.objects.get(cid=cmp)
            if deleted:
                if int(bill.bill_number) > int(deleted.bill_number):
                    deleted.bill_number = bill.bill_number
                    deleted.save()
        else:
            deleted = DeletedPurchases(cid=cmp, bill_number=bill.bill_number)
            deleted.save()

        bill.delete()
        return Response(
            {
                "status": True,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def purchaseBillPdf(request):
    try:
        id = request.GET["Id"]
        purchaseId = request.GET["purchaseId"]

        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        bill = Purchases.objects.get(cid=cmp, bill_no=purchaseId)
        items = Purchase_items.objects.filter(cid=cmp, pid=bill)

        total = bill.total_amount
        words_total = num2words(total)

        context = {"bill": bill, "cmp": cmp, "items": items, "total": words_total}

        template_path = "purchase_bill_pdf.html"
        fname = "PurchaseBill_" + str(bill.bill_number)
        # Create a Django response object, and specify content_type as pdftemp_
        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = f"attachment; filename = {fname}.pdf"
        # find the template and render it.
        template = get_template(template_path)
        html = template.render(context)

        # create a pdf
        pisa_status = pisa.CreatePDF(html, dest=response)
        # if error then show some funny view
        if pisa_status.err:
            return HttpResponse("We had some errors <pre>" + html + "</pre>")
        return response
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def sharePurchaseBillToEmail(request):
    try:
        id = request.data["Id"]
        purchaseId = request.data["purchaseId"]

        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        bill = Purchases.objects.get(cid=cmp, bill_no=purchaseId)
        items = Purchase_items.objects.filter(cid=cmp, pid=bill)

        total = bill.total_amount
        words_total = num2words(total)

        emails_string = request.data["email_ids"]

        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(",")]
        email_message = request.data["email_message"]
        # print(emails_list)

        context = {"bill": bill, "cmp": cmp, "items": items, "total": words_total}

        template_path = "purchase_bill_pdf.html"
        template = get_template(template_path)

        html = template.render(context)
        result = BytesIO()
        pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
        pdf = result.getvalue()
        filename = f"PurchaseBill_{bill.bill_number}.pdf"
        subject = f"PurchaseBill_{bill.bill_number}"
        email = EmailMessage(
            subject,
            f"Hi,\nPlease find the attached details - PURCHASE BILL-{bill.bill_number}. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.phone_number}",
            from_email=settings.EMAIL_HOST_USER,
            to=emails_list,
        )
        email.attach(filename, pdf, "application/pdf")
        email.send(fail_silently=False)

        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("PUT",))
def updatePurchase(request):
    try:
        usr = User.objects.get(id=request.data["Id"])
        cmp = Company.objects.get(user=usr)
        purchase = Purchases.objects.get(bill_no=request.data["purchaseId"])

        mutable_data = deepcopy(request.data)
        mutable_data["cid"] = cmp.cmp_id

        if request.data["party"] == "false":
            mutable_data["party_name"] = ""
            mutable_data["phone_number"] = ""
            mutable_data["gstin"] = ""

        serializer = PurchaseSerializer(purchase, data=mutable_data)
        if serializer.is_valid():
            serializer.save()

            purItems = json.loads(request.data["salesItems"])

            # delete old transactions and stock updates
            items = Purchase_items.objects.filter(pid=purchase)
            for i in items:
                itm = Items.objects.get(id=i.item.id)
                itm.stock -= i.quantity
                itm.save()
                Item_transactions.objects.filter(
                    bill_number=purchase.bill_number, type="Purchase", item=itm
                ).delete()
            Purchase_items.objects.filter(pid=purchase).delete()

            # Add sales items and corresponding item transactions
            for ele in purItems:
                itm = Items.objects.get(id=int(ele.get("item")))
                qty = int(ele.get("quantity"))
                hsn = ele.get("hsn")
                price = ele.get("price")
                tax = (
                    ele.get("taxGst")
                    if request.data["state_of_supply"] == "State"
                    else ele.get("taxIgst")
                )

                Purchase_items.objects.create(
                    cid=cmp,
                    pid=purchase,
                    item=itm,
                    name=itm.name,
                    hsn=hsn,
                    quantity=qty,
                    rate=float(price),
                    tax=tax,
                    total=float(ele.get("total")),
                )

            # Add sales details in items transactions
            for itm in purItems:
                tItem = Items.objects.get(id=itm.get("item"), cid=cmp)
                Item_transactions.objects.create(
                    cid=cmp,
                    item=tItem,
                    type="Purchase",
                    date=purchase.date,
                    quantity=itm.get("quantity"),
                    bill_number=purchase.bill_number,
                )
                tItem.stock += int(itm.get("quantity"))
                tItem.save()

            return Response(
                {"status": True, "data": serializer.data}, status=status.HTTP_200_OK
            )
        else:
            return Response(
                {"status": False, "data": serializer.errors},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


# Stock Reports
@api_view(("GET",))
def getStockReports(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        stockList = []
        items = Items.objects.filter(cid=cmp)

        for item in items:
            stockIn = 0
            stockOut = 0
            for i in Item_transactions.objects.filter(cid=cmp, item=item.id).filter(
                type="Purchase"
            ):
                stockIn += i.quantity

            for i in Item_transactions.objects.filter(cid=cmp, item=item.id).filter(
                type="Sale"
            ):
                stockOut += i.quantity

            dict = {
                "name": item.name,
                "stockIn": stockIn,
                "stockOut": stockOut,
                "balance": item.stock,
            }

            stockList.append(dict)

        return Response(
            {
                "status": True,
                "items": ItemSerializer(items, many=True).data,
                "stock": stockList,
                "count": items.count(),
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getItemStockReports(request):
    try:
        usr = User.objects.get(id=request.GET["Id"])
        cmp = Company.objects.get(user=usr)

        stockList = []
        item = Items.objects.get(cid=cmp, id=request.GET["itemId"])

        stockIn = 0
        stockOut = 0
        for i in Item_transactions.objects.filter(cid=cmp, item=item).filter(
            type="Purchase"
        ):
            stockIn += i.quantity

        for i in Item_transactions.objects.filter(cid=cmp, item=item).filter(
            type="Sale"
        ):
            stockOut += i.quantity

        dict = {
            "name": item.name,
            "stockIn": stockIn,
            "stockOut": stockOut,
            "balance": item.stock,
        }

        stockList.append(dict)

        return Response(
            {
                "status": True,
                "stock": stockList,
                "balance": item.stock,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def shareStockReportsToEmail(request):
    try:
        id = request.data["Id"]

        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        emails_string = request.data["email_ids"]
        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(",")]
        email_message = request.data["email_message"]

        excelfile = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title="Stock Reports", index=1)

        stockList = []
        items = Items.objects.filter(cid=cmp)

        for item in items:
            stockIn = 0
            stockOut = 0
            for i in Item_transactions.objects.filter(cid=cmp, item=item.id).filter(
                type="Purchase"
            ):
                stockIn += i.quantity

            for i in Item_transactions.objects.filter(cid=cmp, item=item.id).filter(
                type="Sale"
            ):
                stockOut += i.quantity

            dict = {
                "name": item.name,
                "stockIn": stockIn,
                "stockOut": stockOut,
                "balance": item.stock,
            }
            stockList.append(dict)

        columns = ["#", "Item", "Stock In", "Stock Out", "Balance"]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=False
            )
            cell.font = Font(bold=True)

        # Iterate through all coins
        sl_no = 0
        for _, bill in enumerate(stockList, 1):
            row_num += 1
            sl_no += 1
            # Define the data for each cell in the row
            name, stockin, stockout, bal = (
                bill.get(key) for key in ["name", "stockIn", "stockOut", "balance"]
            )
            row = [
                sl_no,
                name,
                stockin,
                stockout,
                bal,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.protection = Protection(locked=True)
        workbook.save(excelfile)

        mail_subject = f"Stock Reports - {date.today()}"
        message = f"Hi,\nPlease find the STOCK REPORTS file attached. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.phone_number}"
        message = EmailMessage(
            mail_subject, message, settings.EMAIL_HOST_USER, emails_list
        )
        message.attach(
            f"Stock Reports-{date.today()}.xlsx",
            excelfile.getvalue(),
            "application/vnd.ms-excel",
        )
        message.send(fail_silently=False)

        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def shareSalesReportsToEmail(request):
    try:
        id = request.data["Id"]

        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        emails_string = request.data["email_ids"]
        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(",")]
        email_message = request.data["email_message"]

        excelfile = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title="Sales Reports", index=1)

        stockList = []
        items = Sales.objects.filter(cid=cmp)

        for item in items:
            dict = {
                "date": item.date,
                "invoice_no": item.bill_number,
                "name": item.party_name,
                "amount": item.total_amount,
            }
            stockList.append(dict)

        columns = ["#", "Date", "Invoice No", "Party Name", "Amount"]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=False
            )
            cell.font = Font(bold=True)

        # Iterate through all coins
        sl_no = 0
        for _, bill in enumerate(stockList, 1):
            row_num += 1
            sl_no += 1
            # Define the data for each cell in the row
            date, invoice_no, name, amount = (
                bill.get(key) for key in ["date", "invoice_no", "name", "amount"]
            )
            row = [
                sl_no,
                date,
                invoice_no,
                name,
                amount,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.protection = Protection(locked=True)
        workbook.save(excelfile)

        mail_subject = "Sales Reports"
        message = f"Hi,\nPlease find the SALES REPORTS file attached. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.phone_number}"
        message = EmailMessage(
            mail_subject, message, settings.EMAIL_HOST_USER, emails_list
        )
        message.attach(
            f"Sales Reports.xlsx",
            excelfile.getvalue(),
            "application/vnd.ms-excel",
        )
        message.send(fail_silently=False)

        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getSalesReportDetails(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        sales = Sales.objects.filter(cid=cmp)

        current_year = datetime.now().year

        monthly_sales_data = defaultdict(int)
        for month in range(1, 13):
            monthly_sales_data[month] = (
                Sales.objects.filter(
                    date__month=month, date__year=current_year, cid=cmp
                ).aggregate(total_sales=Sum("total_amount"))["total_sales"]
                or 0
            )

        # Retrieve yearly sales data
        yearly_sales_data = defaultdict(int)
        for year in range(2022, current_year + 1):
            yearly_sales_data[year] = (
                Sales.objects.filter(date__year=year, cid=cmp).aggregate(
                    total_sales=Sum("total_amount")
                )["total_sales"]
                or 0
            )

        # Prepare data for chart
        month_names = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]
        monthly_labels = [
            f"{month_names[month - 1]} {current_year}" for month in range(1, 13)
        ]
        monthly_sales = [monthly_sales_data[month] for month in range(1, 13)]

        yearly_labels = [str(year) for year in range(2014, current_year + 1)]
        yearly_sales = [
            yearly_sales_data[year] for year in range(2014, current_year + 1)
        ]

        # Prepare data for chart
        chart_data = {
            "monthly_labels": monthly_labels,
            "monthly_sales": monthly_sales,
            "yearly_labels": yearly_labels,
            "yearly_sales": yearly_sales,
        }

        serializer = SalesSerializer(sales, many=True)
        return Response(
            {
                "status": True,
                "company": CompanySerializer(cmp).data,
                "sales": serializer.data,
                "chart": chart_data,
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def sharePurchaseReportsToEmail(request):
    try:
        id = request.data["Id"]

        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        emails_string = request.data["email_ids"]
        # Split the string by commas and remove any leading or trailing whitespace
        emails_list = [email.strip() for email in emails_string.split(",")]
        email_message = request.data["email_message"]

        excelfile = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        worksheet = workbook.create_sheet(title="Purchase Reports", index=1)

        stockList = []
        items = Purchases.objects.filter(cid=cmp)

        for item in items:
            dict = {
                "date": item.date,
                "invoice_no": item.bill_number,
                "name": item.party_name,
                "amount": item.total_amount,
            }
            stockList.append(dict)

        columns = ["#", "Date", "Invoice No", "Party Name", "Amount"]
        row_num = 1

        # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(columns, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = column_title
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=False
            )
            cell.font = Font(bold=True)

        # Iterate through all coins
        sl_no = 0
        for _, bill in enumerate(stockList, 1):
            row_num += 1
            sl_no += 1
            # Define the data for each cell in the row
            date, invoice_no, name, amount = (
                bill.get(key) for key in ["date", "invoice_no", "name", "amount"]
            )
            row = [
                sl_no,
                date,
                invoice_no,
                name,
                amount,
            ]

            # Assign the data for each cell of the row
            for col_num, cell_value in enumerate(row, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = cell_value
                cell.protection = Protection(locked=True)
        workbook.save(excelfile)

        mail_subject = "Purchase Reports"
        message = f"Hi,\nPlease find the PURCHASE REPORTS file attached. \n{email_message}\n\n--\nRegards,\n{cmp.company_name}\n{cmp.address}\n{cmp.state} - {cmp.country}\n{cmp.phone_number}"
        message = EmailMessage(
            mail_subject, message, settings.EMAIL_HOST_USER, emails_list
        )
        message.attach(
            f"Purchase Reports.xlsx",
            excelfile.getvalue(),
            "application/vnd.ms-excel",
        )
        message.send(fail_silently=False)

        return Response({"status": True}, status=status.HTTP_200_OK)
    except Exception as e:
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("GET",))
def getPurchaseReportDetails(request, id):
    try:
        usr = User.objects.get(id=id)
        cmp = Company.objects.get(user=usr)

        purchases = Purchases.objects.filter(cid=cmp)
        total_purchase_amount = purchases.aggregate(total_amount=Sum("total_amount"))[
            "total_amount"
        ]

        current_year = datetime.now().year

        monthly_purchase_data = defaultdict(int)
        for month in range(1, 13):
            monthly_purchase_data[month] = (
                Purchases.objects.filter(
                    date__month=month, date__year=current_year, cid=cmp
                ).aggregate(total_sales=Sum("total_amount"))["total_sales"]
                or 0
            )

        # Retrieve yearly sales data
        yearly_purchases_data = defaultdict(int)
        for year in range(2022, current_year + 1):
            yearly_purchases_data[year] = (
                Purchases.objects.filter(date__year=year, cid=cmp).aggregate(
                    total_sales=Sum("total_amount")
                )["total_sales"]
                or 0
            )

        # Prepare data for chart
        month_names = [
            "Jan",
            "Feb",
            "Mar",
            "Apr",
            "May",
            "Jun",
            "Jul",
            "Aug",
            "Sep",
            "Oct",
            "Nov",
            "Dec",
        ]
        monthly_labels = [
            f"{month_names[month - 1]} {current_year}" for month in range(1, 13)
        ]
        monthly_purchases = [monthly_purchase_data[month] for month in range(1, 13)]

        yearly_labels = [str(year) for year in range(2014, current_year + 1)]
        yearly_purchases = [
            yearly_purchases_data[year] for year in range(2014, current_year + 1)
        ]

        # Prepare data for chart
        chart_data = {
            "monthly_labels": monthly_labels,
            "monthly_purchases": monthly_purchases,
            "yearly_labels": yearly_labels,
            "yearly_purchases": yearly_purchases,
        }

        serializer = PurchaseSerializer(purchases, many=True)
        return Response(
            {
                "status": True,
                "company": CompanySerializer(cmp).data,
                "purchases": serializer.data,
                "chart": chart_data,
                "total_purchase_amount": f"{total_purchase_amount:.2f}",
            },
            status=status.HTTP_200_OK,
        )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )


@api_view(("POST",))
def forgotPassword(request):
    try:
        email = request.data["email"]
        if User.objects.filter(email=email).exists():
            user = User.objects.filter(email=email).first()
            password = str(randint(100000, 999999))
            # print(password)
            user.set_password(password)
            user.save()

            # SEND MAIL CODE
            subject = "Forgot Password"
            message = f"Dear user,\nYour Password has been reset as you requested. You can login with the password given below\n\nPassword:{password}"
            recipient = user.email
            send_mail(subject, message, settings.EMAIL_HOST_USER, [recipient])

            return Response(
                {"status": True},
                status=status.HTTP_200_OK,
            )
        else:
            return Response(
                {"status": False, "message": "User does not exists..!"},
                status=status.HTTP_400_BAD_REQUEST,
            )
    except Exception as e:
        print(e)
        return Response(
            {"status": False, "message": str(e)},
            status=status.HTTP_500_INTERNAL_SERVER_ERROR,
        )
